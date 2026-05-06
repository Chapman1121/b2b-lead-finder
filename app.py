"""
B2B Lead Finder – Streamlit App
================================
Run:  streamlit run app.py
"""

import sys
import uuid
import json
import subprocess
import datetime
import time
import pandas as pd
from pathlib import Path

import streamlit as st

# ─── Config ───────────────────────────────────────────────────────────────────
HISTORY_FILE = Path(__file__).parent / "history.json"
RESULTS_DIR  = Path(__file__).parent / "results"

LOCATIONS = {
    "Singapore": [
        "Singapore",
        "Central Region",
        "East Region",
        "North Region",
        "North-East Region",
        "West Region",
        "Orchard",
        "Marina Bay",
        "Tanjong Pagar",
        "Bugis",
        "Jurong",
        "Tampines",
        "Woodlands",
        "Ang Mo Kio",
        "Bishan",
        "Clementi",
        "Bedok",
        "Pasir Ris",
        "Punggol",
        "Sengkang",
        "Hougang",
        "Serangoon",
        "Buona Vista",
        "Novena",
        "Toa Payoh",
    ],
    "Malaysia": [
        "Malaysia",
        "Kuala Lumpur",
        "Selangor",
        "Johor",
        "Penang",
        "Perak",
        "Sabah",
        "Sarawak",
        "Pahang",
        "Negeri Sembilan",
        "Melaka",
        "Kedah",
        "Kelantan",
        "Terengganu",
        "Perlis",
        "Putrajaya",
        "Labuan",
    ],
}

# ─── History ──────────────────────────────────────────────────────────────────
def load_history():
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def save_history(query, country, leads):
    history = load_history()
    entry = {
        "query"    : query,
        "country"  : country,
        "count"    : len(leads),
        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "leads"    : leads,
    }
    for i, h in enumerate(history):
        if h["query"].lower() == query.lower() and h["country"] == country:
            history[i] = entry
            HISTORY_FILE.write_text(json.dumps(history, indent=2, ensure_ascii=False), encoding="utf-8")
            return
    history.insert(0, entry)
    HISTORY_FILE.write_text(json.dumps(history[:30], indent=2, ensure_ascii=False), encoding="utf-8")

# ─── Excel helper ─────────────────────────────────────────────────────────────
@st.cache_data
def to_excel(records):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb  = openpyxl.Workbook(); ws = wb.active; ws.title = "Leads"
    hdrs  = list(records[0].keys())
    hf    = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1A1A2E")
    thin  = Side(style="thin", color="DDDDDD")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font=hf; c.fill=hfill; c.border=bdr
        c.alignment=Alignment(horizontal="center", vertical="center")
    for ri, row in enumerate(records, 2):
        for ci, key in enumerate(hdrs, 1):
            c = ws.cell(row=ri, column=ci, value=row.get(key,""))
            c.border=bdr; c.alignment=Alignment(vertical="center")
            if key=="Google Maps" and row.get(key):
                c.hyperlink=row[key]; c.font=Font(color="4361EE", underline="single")
            elif key=="Name":
                c.font=Font(bold=True)
    for i in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(i)].width = 28
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def read_job(job_id):
    """Read the current state of a job from its JSON file."""
    job_file = RESULTS_DIR / f"{job_id}.json"
    if not job_file.exists():
        return None
    try:
        return json.loads(job_file.read_text(encoding="utf-8"))
    except Exception:
        return None

# ─── Streamlit UI ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="B2B Lead Finder", page_icon="🔍", layout="wide")
st.markdown("""
<style>
  [data-testid="stSidebar"] { background: #161628; }
  .block-container { padding-top: 1.5rem; }
</style>""", unsafe_allow_html=True)

# Session state defaults
for k, v in {
    "leads"     : [],
    "status"    : "",
    "job_id"    : None,
    "running"   : False,
    "proc_pid"  : None,
    "total"     : 0,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🔍 B2B Lead Finder")
    st.caption("Google Maps · Live Scraper")
    st.divider()
    st.markdown("**Previous Searches**")
    history = load_history()
    if not history:
        st.caption("No searches yet")
    else:
        for idx, h in enumerate(history):
            if st.button(f"{h['query']} · {h['country']} · {h['count']} leads",
                         key=f"h_{idx}", use_container_width=True):
                st.session_state.leads   = h["leads"]
                st.session_state.running = False
                st.session_state.job_id  = None
                st.session_state.status  = f"✅ Loaded: **{h['query']}** in {h['country']} — {h['count']} leads"
                st.rerun()

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("## B2B Lead Finder")
st.caption("Find business leads from Google Maps — leads appear live as they're found")

c1, c2, c3, c4, c5 = st.columns([3, 1.5, 2, 1, 1])
with c1:
    query = st.text_input("Search", placeholder="e.g. interior designers, dentists, lawyers …", label_visibility="collapsed")
with c2:
    country = st.selectbox("Country", list(LOCATIONS.keys()), label_visibility="collapsed")
with c3:
    location = st.selectbox("State / Region", LOCATIONS[country], label_visibility="collapsed")
with c4:
    start_btn = st.button("🚀 Start", use_container_width=True, type="primary",
                          disabled=st.session_state.running)
with c5:
    stop_btn = st.button("⏹ Stop", use_container_width=True,
                         disabled=not st.session_state.running)

status_ph   = st.empty()
progress_ph = st.empty()
table_ph    = st.empty()

# ── Stop ──────────────────────────────────────────────────────────────────────
if stop_btn and st.session_state.running:
    # Kill the subprocess if we still have its PID
    if st.session_state.proc_pid:
        try:
            import os, signal
            os.kill(st.session_state.proc_pid, signal.SIGTERM)
        except Exception:
            pass
    st.session_state.running = False
    st.session_state.status  = "⏹ Stopped."
    st.rerun()

# ── Start ─────────────────────────────────────────────────────────────────────
if start_btn and query and not st.session_state.running:
    job_id = str(uuid.uuid4())
    RESULTS_DIR.mkdir(exist_ok=True)

    worker = Path(__file__).parent / "scraper_worker.py"
    proc = subprocess.Popen(
        [sys.executable, str(worker), query, location, job_id],
        # No stdout/stderr pipes — let it run truly independently
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

    st.session_state.leads    = []
    st.session_state.job_id   = job_id
    st.session_state.running  = True
    st.session_state.proc_pid = proc.pid
    st.session_state.status   = "🌐 Starting scraper …"
    st.session_state.total    = 80
    st.rerun()

# ── Poll running job ──────────────────────────────────────────────────────────
if st.session_state.running and st.session_state.job_id:
    job_id = st.session_state.job_id
    data   = read_job(job_id)

    if data is None:
        # File not written yet — still starting up
        status_ph.info("🌐 Starting scraper …")
        time.sleep(1)
        st.rerun()
    else:
        leads   = data.get("leads", [])
        message = data.get("message", "")
        total   = data.get("total", st.session_state.total) or 80
        done    = data.get("status") == "done"

        st.session_state.leads = leads
        st.session_state.total = total

        n   = len(leads)
        pct = min(n / total, 1.0) if total else 0

        if done:
            st.session_state.running = False
            st.session_state.job_id  = None
            progress_ph.empty()
            status_ph.success(f"✅ Done! Found **{n}** leads for **{query}** in **{location}**")
            save_history(query, location, leads)
        else:
            status_ph.info(message or f"⚡ Scraping … **{n}** leads found so far")
            if n > 0:
                progress_ph.progress(pct, text=f"{n} / {total} processed")

        if leads:
            table_ph.dataframe(
                pd.DataFrame(leads),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Google Maps": st.column_config.LinkColumn("Maps 📍"),
                    "Website"    : st.column_config.LinkColumn("Website 🌐"),
                    "Rating"     : st.column_config.NumberColumn("Rating ⭐", format="%.1f"),
                    "Reviews"    : st.column_config.NumberColumn("Reviews 💬"),
                },
            )

        if not done:
            time.sleep(2)
            st.rerun()

# ── Display existing leads (not running) ─────────────────────────────────────
elif st.session_state.leads and not st.session_state.running:
    if st.session_state.status:
        status_ph.info(st.session_state.status)

    df = pd.DataFrame(st.session_state.leads)

    f1, f2, f3 = st.columns([3, 2, 2])
    with f1:
        sf = st.text_input("Filter", placeholder="Search name, address, phone …", label_visibility="collapsed")
    with f2:
        cf = st.selectbox("Contact", ["All leads", "Has phone", "Has website", "Phone + website"], label_visibility="collapsed")
    with f3:
        so = st.selectbox("Sort", ["Reviews ↓", "Reviews ↑", "Rating ↓", "Name A–Z"], label_visibility="collapsed")

    if sf:
        mask = df.apply(lambda r: sf.lower() in r.astype(str).str.lower().str.cat(), axis=1)
        df   = df[mask]
    if cf == "Has phone":
        df = df[df["Phone"].astype(str).str.len() > 2]
    elif cf == "Has website":
        df = df[df["Website"].astype(str).str.len() > 4]
    elif cf == "Phone + website":
        df = df[(df["Phone"].astype(str).str.len() > 2) & (df["Website"].astype(str).str.len() > 4)]

    sc, asc = {"Reviews ↓":("Reviews",False),"Reviews ↑":("Reviews",True),"Rating ↓":("Rating",False),"Name A–Z":("Name",True)}[so]
    try:
        df = df.sort_values(sc, ascending=asc)
    except Exception:
        pass

    st.caption(f"Showing **{len(df)}** of **{len(st.session_state.leads)}** leads")
    st.dataframe(
        df, use_container_width=True, hide_index=True,
        column_config={
            "Google Maps": st.column_config.LinkColumn("Maps 📍"),
            "Website"    : st.column_config.LinkColumn("Website 🌐"),
            "Rating"     : st.column_config.NumberColumn("Rating ⭐", format="%.1f ⭐"),
            "Reviews"    : st.column_config.NumberColumn("Reviews 💬"),
        },
    )
    st.download_button(
        "⬇️ Download Excel",
        data=to_excel(tuple(st.session_state.leads)),
        file_name=f"leads_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

else:
    st.markdown("""
    <div style="text-align:center;padding:60px 0;color:#7c829e">
      <div style="font-size:3rem">🏢</div>
      <div style="margin-top:12px;font-size:1.1rem">Search for any type of business in any country</div>
      <div style="margin-top:8px;font-size:.85rem">Leads appear live as they're scraped from Google Maps</div>
    </div>""", unsafe_allow_html=True)
